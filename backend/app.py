import csv
import json
import os
import sqlite3
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile

from flask import Flask, jsonify, request
from flask_cors import CORS
from werkzeug.utils import secure_filename

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


app = Flask(__name__)
CORS(app)

DATA_ROOT = Path(os.environ.get("DATA_ROOT", "/data"))
DB_PATH = DATA_ROOT / "zomba.db"
UPLOAD_ROOT = DATA_ROOT / "uploads"
RAW_UPLOAD_ROOT = UPLOAD_ROOT / "raw"

SUPPORTED_EXTENSIONS = {
    "education": {".csv", ".xlsx", ".xls", ".json"},
    "health": {".csv", ".xlsx", ".xls", ".json"},
    "disaster": {".csv", ".json", ".zip", ".shp"},
    "welfare": {".csv", ".xlsx", ".xls", ".json"},
}

DEPARTMENT_SCHEMAS = {
    "education": {
        "entity": "schools",
        "description": "School rosters, enrollment, staffing and facility coordinates.",
        "accepted_formats": [".csv", ".xlsx", ".json"],
        "canonical_fields": ["name", "ward", "lat", "lng", "enrollment", "teachers", "type"],
    },
    "health": {
        "entity": "health_facilities",
        "description": "Facility master lists, service volumes and bed capacity.",
        "accepted_formats": [".csv", ".xlsx", ".json"],
        "canonical_fields": ["name", "ward", "lat", "lng", "type", "beds", "monthly_visits"],
    },
    "disaster": {
        "entity": "disaster_events",
        "description": "Incident records, hazard footprints and affected household estimates.",
        "accepted_formats": [".csv", ".json", ".zip", ".shp"],
        "canonical_fields": ["event_name", "incident_type", "ward", "lat", "lng", "severity", "households_affected", "reported_on"],
    },
    "welfare": {
        "entity": "welfare_records",
        "description": "Social support programs, beneficiaries and payment cycles.",
        "accepted_formats": [".csv", ".xlsx", ".json"],
        "canonical_fields": ["program_name", "ward", "beneficiaries", "amount", "status", "reporting_period"],
    },
}

FIELD_ALIASES = {
    "name": ["name", "school_name", "facility_name", "institution", "institution_name"],
    "ward": ["ward", "ward_name", "ta", "traditional_authority", "area"],
    "lat": ["lat", "latitude", "y", "gps_lat"],
    "lng": ["lng", "lon", "long", "longitude", "x", "gps_lng"],
    "enrollment": ["enrollment", "students", "student_enrollment", "pupils"],
    "teachers": ["teachers", "teacher_count", "staff_teachers"],
    "type": ["type", "facility_type", "school_type", "category"],
    "beds": ["beds", "bed_count", "capacity_beds"],
    "monthly_visits": ["monthly_visits", "visits", "monthly_attendance", "clients_per_month"],
    "event_name": ["event_name", "name", "incident_name"],
    "incident_type": ["incident_type", "hazard_type", "disaster_type", "type"],
    "severity": ["severity", "risk_level", "alert_level"],
    "households_affected": ["households_affected", "affected_households", "households"],
    "reported_on": ["reported_on", "date_reported", "report_date", "incident_date"],
    "program_name": ["program_name", "programme", "program", "intervention"],
    "beneficiaries": ["beneficiaries", "households_supported", "people_supported"],
    "amount": ["amount", "transfer_amount", "funding_amount", "value"],
    "status": ["status", "case_status", "delivery_status"],
    "reporting_period": ["reporting_period", "period", "month", "quarter"],
}

NUMERIC_FIELDS = {
    "lat",
    "lng",
    "enrollment",
    "teachers",
    "beds",
    "monthly_visits",
    "households_affected",
    "beneficiaries",
    "amount",
}


def get_db():
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def normalize_key(value):
    return "".join(ch.lower() if ch.isalnum() else "_" for ch in str(value).strip()).strip("_")


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def canonicalize_record(record):
    return {normalize_key(key): normalize_value(value) for key, value in record.items() if normalize_key(key)}


def parse_numeric(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    cleaned = str(value).replace(",", "").strip()
    if not cleaned:
        return None
    try:
        number = float(cleaned)
    except ValueError:
        return value
    return int(number) if number.is_integer() else number


def harmonize_record(record, department):
    schema = DEPARTMENT_SCHEMAS[department]
    harmonized = {}

    for canonical_field in schema["canonical_fields"]:
        value = None
        for alias in FIELD_ALIASES.get(canonical_field, [canonical_field]):
            alias_value = record.get(normalize_key(alias))
            if alias_value not in (None, ""):
                value = alias_value
                break

        if canonical_field in NUMERIC_FIELDS:
            value = parse_numeric(value)
        elif isinstance(value, str):
            value = value.strip()

        harmonized[canonical_field] = value

    extras = {key: value for key, value in record.items() if key not in harmonized and value not in (None, "")}
    if extras:
        harmonized["_extra"] = extras

    return harmonized


def load_csv_records(path):
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [canonicalize_record(row) for row in csv.DictReader(handle)]


def load_json_records(path):
    with path.open("r", encoding="utf-8") as handle:
        payload = json.load(handle)

    if isinstance(payload, dict):
        if isinstance(payload.get("records"), list):
            payload = payload["records"]
        else:
            payload = [payload]

    return [canonicalize_record(row) for row in payload if isinstance(row, dict)]


def load_excel_records(path):
    if load_workbook is None:
        raise ValueError("Excel support is unavailable because openpyxl is not installed.")

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()

    if not rows:
        return []

    header = [normalize_key(cell) for cell in rows[0] if cell is not None]
    records = []
    for row in rows[1:]:
        values = list(row[: len(header)])
        record = {header[index]: normalize_value(values[index]) if index < len(values) else None for index in range(len(header))}
        records.append(record)
    return records


def load_shapefile_metadata(path):
    metadata = {
        "source_file": path.name,
        "stored_as": str(path),
        "geometry_type": "shapefile_bundle" if path.suffix.lower() == ".zip" else "shapefile_component",
    }

    if path.suffix.lower() == ".zip":
        with ZipFile(path) as archive:
            members = archive.namelist()
        metadata["archive_members"] = members

    return [canonicalize_record(metadata)]


def load_records(path, department):
    extension = path.suffix.lower()

    if extension == ".csv":
        return load_csv_records(path)
    if extension == ".json":
        return load_json_records(path)
    if extension in {".xlsx", ".xls"}:
        return load_excel_records(path)
    if department == "disaster" and extension in {".zip", ".shp"}:
        return load_shapefile_metadata(path)

    raise ValueError(f"Unsupported file type: {extension}")


def clean_records(records):
    cleaned = []
    dropped = 0

    for record in records:
        normalized = {
            normalize_key(key): normalize_value(value)
            for key, value in record.items()
            if normalize_key(key)
        }
        if any(value not in (None, "") for value in normalized.values()):
            cleaned.append(normalized)
        else:
            dropped += 1

    return cleaned, {"dropped_empty_rows": dropped}


def summarize_schema(records):
    keys = set()
    for record in records[:50]:
        keys.update(record.keys())
    return sorted(keys)


def create_pipeline_job(conn, department, source_format, original_filename, stored_filename, stored_path):
    cursor = conn.execute(
        """
        INSERT INTO pipeline_jobs (
            department, source_format, original_filename, stored_filename, stored_path,
            status, pipeline_stage, message, started_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            department,
            source_format,
            original_filename,
            stored_filename,
            stored_path,
            "running",
            "load",
            "File received and queued for processing.",
            datetime.utcnow().isoformat(),
        ),
    )
    conn.commit()
    return cursor.lastrowid


def update_pipeline_job(conn, job_id, **fields):
    if not fields:
        return

    assignments = ", ".join(f"{field} = ?" for field in fields)
    values = list(fields.values()) + [job_id]
    conn.execute(f"UPDATE pipeline_jobs SET {assignments} WHERE id = ?", values)
    conn.commit()


def save_staged_dataset(conn, job_id, department, source_format, original_filename, raw_path, cleaned_records, harmonized_records, quality_report):
    conn.execute(
        """
        INSERT INTO staged_datasets (
            job_id, department, source_format, original_filename, raw_path,
            record_count, schema_json, cleaned_preview_json, harmonized_preview_json,
            quality_report_json, status, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            job_id,
            department,
            source_format,
            original_filename,
            raw_path,
            len(harmonized_records),
            json.dumps(summarize_schema(cleaned_records)),
            json.dumps(cleaned_records[:5]),
            json.dumps(harmonized_records[:5]),
            json.dumps(quality_report),
            "ready_for_integration",
            datetime.utcnow().isoformat(),
        ),
    )
    conn.commit()


def process_upload(job_id, department, source_format, original_filename, stored_path):
    conn = get_db()

    try:
        raw_records = load_records(Path(stored_path), department)
        update_pipeline_job(
            conn,
            job_id,
            pipeline_stage="clean",
            message=f"Loaded {len(raw_records)} raw records.",
        )

        cleaned_records, clean_report = clean_records(raw_records)
        update_pipeline_job(
            conn,
            job_id,
            pipeline_stage="harmonize",
            message=f"Cleaned {len(cleaned_records)} records; harmonizing schema.",
        )

        harmonized_records = [harmonize_record(record, department) for record in cleaned_records]
        quality_report = {
            "raw_records": len(raw_records),
            "clean_records": len(cleaned_records),
            "dropped_empty_rows": clean_report["dropped_empty_rows"],
            "missing_ward": sum(1 for record in harmonized_records if not record.get("ward")),
            "missing_coordinates": sum(
                1 for record in harmonized_records if "lat" in record and (record.get("lat") is None or record.get("lng") is None)
            ),
        }

        save_staged_dataset(
            conn,
            job_id,
            department,
            source_format,
            original_filename,
            stored_path,
            cleaned_records,
            harmonized_records,
            quality_report,
        )
        update_pipeline_job(
            conn,
            job_id,
            status="ready",
            pipeline_stage="integrate",
            message="Pipeline completed through harmonization. Dataset is ready for central DB integration.",
            record_count=len(harmonized_records),
            finished_at=datetime.utcnow().isoformat(),
        )
    except Exception as exc:
        update_pipeline_job(
            conn,
            job_id,
            status="failed",
            pipeline_stage="load",
            message=str(exc),
            finished_at=datetime.utcnow().isoformat(),
        )
        raise
    finally:
        conn.close()


def seed_demo_data(conn):
    schools_count = conn.execute("SELECT COUNT(*) AS count FROM schools").fetchone()["count"]
    if schools_count:
        return

    schools = [
        ("Zomba Primary School", "Zomba Central", -15.3833, 35.3167, 850, 18, "Primary"),
        ("Naisi Primary School", "Naisi", -15.4012, 35.2890, 620, 14, "Primary"),
        ("Domasi Secondary School", "Domasi", -15.2741, 35.4123, 430, 22, "Secondary"),
        ("Likangala Primary School", "Likangala", -15.3560, 35.2650, 710, 16, "Primary"),
        ("Malemia Primary School", "Malemia", -15.4234, 35.3456, 540, 12, "Primary"),
        ("Mpemba Primary School", "Mpemba", -15.4567, 35.2345, 390, 9, "Primary"),
        ("Chikanda Secondary School", "Chikanda", -15.3100, 35.3890, 280, 18, "Secondary"),
        ("Machinjiri Primary School", "Machinjiri", -15.3678, 35.2123, 960, 21, "Primary"),
        ("Songani Primary School", "Songani", -15.4890, 35.3210, 470, 11, "Primary"),
        ("Thondwe Secondary School", "Thondwe", -15.3234, 35.4560, 380, 20, "Secondary"),
        ("Chingwe Primary School", "Chingwe", -15.2980, 35.2780, 510, 13, "Primary"),
        ("Chiradzulu Road Primary", "Zomba Central", -15.3750, 35.3300, 720, 17, "Primary"),
    ]

    health_facilities = [
        ("Zomba Central Hospital", "Zomba Central", -15.3850, 35.3200, "Hospital", 400, 3200),
        ("Naisi Health Centre", "Naisi", -15.4020, 35.2900, "Health Centre", 20, 480),
        ("Domasi Community Hospital", "Domasi", -15.2750, 35.4130, "Hospital", 80, 920),
        ("Likangala Health Post", "Likangala", -15.3570, 35.2660, "Health Post", 0, 210),
        ("Malemia Dispensary", "Malemia", -15.4240, 35.3460, "Dispensary", 5, 180),
        ("Mpemba Health Centre", "Mpemba", -15.4570, 35.2350, "Health Centre", 15, 340),
        ("Songani Dispensary", "Songani", -15.4890, 35.3220, "Dispensary", 4, 150),
        ("Thondwe Health Centre", "Thondwe", -15.3240, 35.4570, "Health Centre", 18, 390),
        ("Machinjiri Health Post", "Machinjiri", -15.3680, 35.2130, "Health Post", 0, 175),
        ("Chikanda Dispensary", "Chikanda", -15.3110, 35.3900, "Dispensary", 6, 200),
    ]

    population = [
        ("Zomba Central", 45200, 12.4, 9040, 4520, 8136, 32544),
        ("Naisi", 28400, 34.7, 5680, 2840, 5112, 20448),
        ("Domasi", 31200, 28.9, 6240, 3120, 5616, 22464),
        ("Likangala", 19800, 41.2, 3960, 1980, 3564, 14256),
        ("Malemia", 22100, 38.6, 4420, 2210, 3978, 15912),
        ("Mpemba", 17600, 52.3, 3520, 1760, 3168, 12672),
        ("Chikanda", 24300, 29.8, 4860, 2430, 4374, 17496),
        ("Machinjiri", 33800, 22.1, 6760, 3380, 6084, 24336),
        ("Songani", 15400, 44.1, 3080, 1540, 2772, 11088),
        ("Thondwe", 21600, 31.5, 4320, 2160, 3888, 15552),
        ("Chingwe", 18900, 37.8, 3780, 1890, 3402, 13608),
    ]

    conn.executemany(
        "INSERT INTO schools (name, ward, lat, lng, enrollment, teachers, type) VALUES (?, ?, ?, ?, ?, ?, ?)",
        schools,
    )
    conn.executemany(
        "INSERT INTO health_facilities (name, ward, lat, lng, type, beds, monthly_visits) VALUES (?, ?, ?, ?, ?, ?, ?)",
        health_facilities,
    )
    conn.executemany(
        "INSERT INTO population (ward, population, area_km2, households, under5, school_age, adults) VALUES (?, ?, ?, ?, ?, ?, ?)",
        population,
    )
    conn.commit()


def init_db():
    RAW_UPLOAD_ROOT.mkdir(parents=True, exist_ok=True)
    conn = get_db()

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS schools (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT, ward TEXT,
            lat REAL, lng REAL,
            enrollment INTEGER, teachers INTEGER, type TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS health_facilities (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT, ward TEXT,
            lat REAL, lng REAL,
            type TEXT, beds INTEGER, monthly_visits INTEGER
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS population (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ward TEXT, population INTEGER,
            area_km2 REAL, households INTEGER,
            under5 INTEGER, school_age INTEGER, adults INTEGER
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS pipeline_jobs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department TEXT NOT NULL,
            source_format TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            stored_filename TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            status TEXT NOT NULL,
            pipeline_stage TEXT NOT NULL,
            message TEXT,
            record_count INTEGER DEFAULT 0,
            started_at TEXT,
            finished_at TEXT
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS staged_datasets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            job_id INTEGER NOT NULL,
            department TEXT NOT NULL,
            source_format TEXT NOT NULL,
            original_filename TEXT NOT NULL,
            raw_path TEXT NOT NULL,
            record_count INTEGER DEFAULT 0,
            schema_json TEXT,
            cleaned_preview_json TEXT,
            harmonized_preview_json TEXT,
            quality_report_json TEXT,
            status TEXT NOT NULL,
            created_at TEXT,
            FOREIGN KEY(job_id) REFERENCES pipeline_jobs(id)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS uploads_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            department TEXT NOT NULL,
            filename TEXT NOT NULL,
            source_format TEXT NOT NULL,
            job_id INTEGER NOT NULL,
            uploaded_at TEXT NOT NULL,
            FOREIGN KEY(job_id) REFERENCES pipeline_jobs(id)
        )
        """
    )
    conn.commit()

    seed_demo_data(conn)
    conn.close()


def serialize_job(row):
    return dict(row)


@app.route("/api/health")
def health():
    return jsonify({"status": "ok", "message": "Zomba Dashboard API running"})


@app.route("/api/schools")
def get_schools():
    conn = get_db()
    rows = conn.execute("SELECT * FROM schools").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/health-facilities")
def get_health_facilities():
    conn = get_db()
    rows = conn.execute("SELECT * FROM health_facilities").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/population")
def get_population():
    conn = get_db()
    rows = conn.execute("SELECT * FROM population").fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/summary")
def get_summary():
    conn = get_db()
    total_pop = conn.execute("SELECT SUM(population) AS total FROM population").fetchone()["total"]
    total_schools = conn.execute("SELECT COUNT(*) AS total FROM schools").fetchone()["total"]
    total_health = conn.execute("SELECT COUNT(*) AS total FROM health_facilities").fetchone()["total"]
    total_enroll = conn.execute("SELECT SUM(enrollment) AS total FROM schools").fetchone()["total"]
    wards = conn.execute("SELECT COUNT(*) AS total FROM population").fetchone()["total"]
    conn.close()
    return jsonify(
        {
            "total_population": total_pop,
            "total_schools": total_schools,
            "total_health_facilities": total_health,
            "total_enrollment": total_enroll,
            "total_wards": wards,
        }
    )


@app.route("/api/ward/<ward_name>")
def get_ward_summary(ward_name):
    conn = get_db()
    schools = conn.execute("SELECT * FROM schools WHERE ward = ?", (ward_name,)).fetchall()
    facilities = conn.execute("SELECT * FROM health_facilities WHERE ward = ?", (ward_name,)).fetchone()
    population = conn.execute("SELECT * FROM population WHERE ward = ?", (ward_name,)).fetchone()
    conn.close()
    return jsonify(
        {
            "ward": ward_name,
            "population": dict(population) if population else {},
            "schools": [dict(row) for row in schools],
            "health_facility": dict(facilities) if facilities else {},
        }
    )


@app.route("/api/wards")
def get_wards():
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT ward FROM population ORDER BY ward").fetchall()
    conn.close()
    return jsonify([row["ward"] for row in rows])


@app.route("/api/admin/pipeline")
def get_pipeline_overview():
    conn = get_db()
    jobs = conn.execute(
        """
        SELECT id, department, source_format, original_filename, status, pipeline_stage,
               message, record_count, started_at, finished_at
        FROM pipeline_jobs
        ORDER BY id DESC
        LIMIT 12
        """
    ).fetchall()
    aggregates = conn.execute(
        """
        SELECT
            COUNT(*) AS total_jobs,
            SUM(CASE WHEN status = 'ready' THEN 1 ELSE 0 END) AS ready_jobs,
            SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) AS failed_jobs,
            SUM(CASE WHEN pipeline_stage = 'integrate' THEN 1 ELSE 0 END) AS waiting_for_integration
        FROM pipeline_jobs
        """
    ).fetchone()
    staged = conn.execute(
        """
        SELECT department, COUNT(*) AS dataset_count, COALESCE(SUM(record_count), 0) AS total_records
        FROM staged_datasets
        GROUP BY department
        ORDER BY department
        """
    ).fetchall()
    conn.close()

    return jsonify(
        {
            "departments": DEPARTMENT_SCHEMAS,
            "supported_extensions": {key: sorted(value) for key, value in SUPPORTED_EXTENSIONS.items()},
            "jobs": [serialize_job(job) for job in jobs],
            "stats": dict(aggregates),
            "staged_datasets": [dict(row) for row in staged],
            "database_plan": {
                "status": "pending",
                "message": "Pipeline is active through harmonization. Central database merge is intentionally the next phase.",
            },
        }
    )


@app.route("/api/admin/staged-datasets")
def get_staged_datasets():
    conn = get_db()
    rows = conn.execute(
        """
        SELECT s.id, s.department, s.source_format, s.original_filename, s.record_count,
               s.schema_json, s.cleaned_preview_json, s.harmonized_preview_json,
               s.quality_report_json, s.status, s.created_at, p.id AS job_id
        FROM staged_datasets s
        JOIN pipeline_jobs p ON p.id = s.job_id
        ORDER BY s.id DESC
        LIMIT 8
        """
    ).fetchall()
    conn.close()

    datasets = []
    for row in rows:
        item = dict(row)
        item["schema"] = json.loads(item.pop("schema_json") or "[]")
        item["cleaned_preview"] = json.loads(item.pop("cleaned_preview_json") or "[]")
        item["harmonized_preview"] = json.loads(item.pop("harmonized_preview_json") or "[]")
        item["quality_report"] = json.loads(item.pop("quality_report_json") or "{}")
        datasets.append(item)

    return jsonify(datasets)


@app.route("/api/admin/upload", methods=["POST"])
def upload_dataset():
    department = request.form.get("department", "").strip().lower()
    upload = request.files.get("file")

    if department not in DEPARTMENT_SCHEMAS:
        return jsonify({"error": "Unsupported department selected."}), 400

    if upload is None or not upload.filename:
        return jsonify({"error": "No file uploaded."}), 400

    original_filename = secure_filename(upload.filename)
    extension = Path(original_filename).suffix.lower()
    if extension not in SUPPORTED_EXTENSIONS[department]:
        return jsonify(
            {
                "error": f"{department.title()} uploads currently accept: {', '.join(sorted(SUPPORTED_EXTENSIONS[department]))}"
            }
        ), 400

    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
    department_dir = RAW_UPLOAD_ROOT / department
    department_dir.mkdir(parents=True, exist_ok=True)
    stored_filename = f"{timestamp}_{original_filename}"
    stored_path = department_dir / stored_filename
    upload.save(stored_path)

    conn = get_db()
    job_id = create_pipeline_job(conn, department, extension, original_filename, stored_filename, str(stored_path))
    conn.execute(
        """
        INSERT INTO uploads_log (department, filename, source_format, job_id, uploaded_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (department, original_filename, extension, job_id, datetime.utcnow().isoformat()),
    )
    conn.commit()
    conn.close()

    try:
        process_upload(job_id, department, extension, original_filename, str(stored_path))
    except Exception as exc:
        return jsonify({"error": str(exc), "job_id": job_id}), 500

    conn = get_db()
    job = conn.execute("SELECT * FROM pipeline_jobs WHERE id = ?", (job_id,)).fetchone()
    staged_dataset = conn.execute(
        """
        SELECT record_count, quality_report_json
        FROM staged_datasets
        WHERE job_id = ?
        ORDER BY id DESC
        LIMIT 1
        """,
        (job_id,),
    ).fetchone()
    conn.close()

    response = dict(job)
    if staged_dataset:
        response["quality_report"] = json.loads(staged_dataset["quality_report_json"] or "{}")
        response["record_count"] = staged_dataset["record_count"]

    return jsonify(response), 201


init_db()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
